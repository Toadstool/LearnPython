from openpyxl import load_workbook
import operator


def calculateAverageForSubject(subject):
    wb = load_workbook('oceny.xlsx')
    active_sheet = wb.active
    s = {}
    for sheet in wb:
        if(sheet.title== subject):
            sum = 0
            count =0
            for row in sheet.iter_rows(min_col=1, max_col=2):
                sum +=row[1].value
                count+=1
            return  sum/count
    return "Subject not found"
#end def




#funkcji calculateAverageForSubject(subject) wyliczającej średnią ocen dla danego przedmiotu
print("Matematyka "+str(calculateAverageForSubject("Matematyka")))
#funkcji calculateAverageForStudent(student) wyliczającej średnią ocen ze wszystkich przedmiotów dla danego studenta
#print(calculateAverageForStudent("Jan Kowalski"))
#funkcji calculateAverageForAllSubjects() wyliczającej średnią ocen wszystkich uczniów dla wszystkich przedmiotów
#funkcji generateStudentRankingForSubject(subject) wyliczającej ranking uczniów w danym przedmiocie. Ranking oczywiście ma być wyznaczany na podstawie ocen, najwyższe oceny na początku
#funkcji generateStudentRankingForAllSubjects() wyliczającej ranking uczniów na podstawie średniej ocen ze wszystkich przedmiotów, najlepsi studenci na początku
