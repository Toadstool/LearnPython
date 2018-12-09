from openpyxl import load_workbook
import operator


def calculateAverageForSubject(subject):
    wb = load_workbook('oceny.xlsx')
    active_sheet = wb.active
    s = {}
    for sheet in wb:
        if(sheet.title== subject):
            sum = 0
            count =0
            for row in sheet.iter_rows(min_col=1, max_col=2):
                sum +=row[1].value
                count+=1
            return  sum/count
    return "Subject not found"
#end def


def calculateAverageForStudent(student):
    wb = load_workbook('oceny.xlsx')
    active_sheet = wb.active
    s = 0
    for sheet in wb:
        for row in sheet.iter_rows(min_col=1, max_col=2):
            if row[0].value == student:
                s +=int(row[1].value);
    return s/3
#end def

def calculateAverageForAllSubjects():
    wb = load_workbook('oceny.xlsx')
    active_sheet = wb.active
    s = 0
    count =0
    for sheet in wb:
        for row in sheet.iter_rows(min_col=1, max_col=2):
                s +=int(row[1].value);
                count+=1
    return s/count
#end def


def generateStudentRankingForSubject(subject):
    wb = load_workbook('oceny.xlsx')
    active_sheet = wb.active
    s = {}
    for sheet in wb:
        if (sheet.title == subject):
            for row in sheet.iter_rows(min_col=1, max_col=2):
                if row[0].value not in s:
                    s[row[0].value] = int(row[1].value)
                #else:
                #    s.update({row[0].value: s[row[0].value] + int(row[1].value)})
    sl = sorted(s.items(), key=operator.itemgetter(1), reverse=True)
    for i in sl:
        print(i)
#end def


def generateStudentRankingForAllSubjects():
    wb = load_workbook('oceny.xlsx')
    active_sheet = wb.active
    s = {}
    count = len(wb.sheetnames)
    for sheet in wb:
        for row in sheet.iter_rows(min_col=1, max_col=2):
            if row[0].value not in s:
                s[row[0].value] = int(row[1].value/count)
            else:
                s.update({row[0].value: s[row[0].value] + int(row[1].value)/count})
    sl = sorted(s.items(), key=operator.itemgetter(1), reverse=True)
    for i in sl:
        print(i)
#end def


#funkcji calculateAverageForSubject(subject) wyliczającej średnią ocen dla danego przedmiotu
print("Matematyka "+str(calculateAverageForSubject("Matematyka")))
#funkcji calculateAverageForStudent(student) wyliczającej średnią ocen ze wszystkich przedmiotów dla danego studenta
print("Jan Kowalski  "+str(calculateAverageForStudent("Jan Kowalski")))
#funkcji calculateAverageForAllSubjects() wyliczającej średnią ocen wszystkich uczniów dla wszystkich przedmiotów
print("AVG  "+str(calculateAverageForAllSubjects()))
#funkcji generateStudentRankingForSubject(subject) wyliczającej ranking uczniów w danym przedmiocie. Ranking oczywiście ma być wyznaczany na podstawie ocen, najwyższe oceny na początku
generateStudentRankingForSubject("Matematyka")
#funkcji generateStudentRankingForAllSubjects() wyliczającej ranking uczniów na podstawie średniej ocen ze wszystkich przedmiotów, najlepsi studenci na początku
print("Ranking")
generateStudentRankingForAllSubjects()
